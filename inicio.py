import tkinter as tk
from tkinter import ttk, messagebox, filedialog, Menu
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import os
import re

# Definição de uma classe para representar os itens do orçamento
class ItemOrcamento:
    def __init__(self, instrumento, resolucao, capacidade, codigo, modelo, fabricante, cliente, manutencao, valor_total, protocolo, status):
        self.instrumento = instrumento
        self.resolucao = resolucao
        self.capacidade = capacidade
        self.codigo = codigo
        self.modelo = modelo
        self.fabricante = fabricante
        self.cliente = cliente
        self.manutencao = manutencao
        self.valor_total = valor_total
        self.protocolo = protocolo
        self.status = status

# Função para criar e salvar a planilha Excel
def criar_planilha_orcamento(itens, nome_arquivo, numero_orcamento, data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orçamento"

    # Adicionar título com número do orçamento e data
    ws.merge_cells('A1:K1')
    ws['A1'] = f"Orçamento {numero_orcamento} - Data: {data}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal="center")

    # Verificar se há algum item com protocolo
    has_protocolo = any(item.protocolo for item in itens)

    # Verificar se há algum item com status válido
    has_status = any(item.status and item.status.lower() != "none" for item in itens)

    cabecalhos = ["Item", "Instrumento", "Resolução (mm)", "Capacidade (mm)", "Código", "Modelo", "Fabricante", "Cliente", "Manutenção", "Valor Unitário"]
    if has_protocolo:
        cabecalhos.insert(7, "Protocolo")
    if has_status:
        cabecalhos.append("Status")
    ws.append(cabecalhos)

    valor_total = 0
    for index, item in enumerate(itens, start=1):
        row = [index, item.instrumento, item.resolucao, item.capacidade, item.codigo, item.modelo, item.fabricante, item.cliente, item.manutencao, item.valor_total]
        if has_protocolo:
            row.insert(7, item.protocolo)
        if has_status and item.status and item.status.lower() != "none":
            row.append(item.status)
        ws.append(row)
        valor_total += float(item.valor_total) if isinstance(item.valor_total, str) else item.valor_total

    # Adicionar o Valor Total no final
    ws.append([])  # Linha vazia para separação
    ws.append(["Valor Total", f"R${valor_total:.2f}"])
    ultima_linha = ws.max_row
    ws.cell(row=ultima_linha, column=9).font = Font(bold=True)

    # Estilizar os cabeçalhos
    for col_num, column_title in enumerate(cabecalhos, 1):
        col_letter = get_column_letter(col_num)
        ws[f"{col_letter}2"].font = Font(bold=True)
        ws[f"{col_letter}2"].alignment = Alignment(horizontal="center")
        # Ajustar a largura da coluna
        ws.column_dimensions[col_letter].width = len(column_title) + 5

    # Ajustar a largura das colunas automaticamente
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    # Ajustar a altura das linhas
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=len(cabecalhos)):
        max_height = 0
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if cell.value:
                text_lines = str(cell.value).count('\n') + 1
                text_height = text_lines * 15  # Estimativa de 15 unidades por linha
                if text_height > max_height:
                    max_height = text_height
        ws.row_dimensions[row[0].row].height = max_height

    # Obter o nome do cliente do primeiro item (assumindo que todos os itens são do mesmo cliente)
    cliente = itens[0].cliente if itens else "Cliente_Desconhecido"
    
    # Sanitize the file name and path
    cliente = re.sub(r'[<>:"/\\|?*]', '', cliente)  # Remove invalid characters
    numero_orcamento = re.sub(r'[<>:"/\\|?*]', '', numero_orcamento)  # Remove invalid characters
    nome_arquivo = re.sub(r'[<>:"/\\|?*]', '', nome_arquivo)  # Remove invalid characters
    
    # Replace forward slashes with hyphens in the date
    data = data.replace('/', '-')
    
    # Create or get the client folder
    pasta_cliente = obter_pasta_cliente(cliente)
    
    # Modify the file name to include the quote number
    nome_arquivo_completo = f"Orcamento_{numero_orcamento}_{nome_arquivo}"
    caminho_completo = os.path.join(pasta_cliente, nome_arquivo_completo)

    # Ensure the directory exists
    os.makedirs(os.path.dirname(caminho_completo), exist_ok=True)

    try:
        wb.save(caminho_completo)
        messagebox.showinfo("Sucesso", f"Planilha '{nome_arquivo_completo}' criada com sucesso na pasta do cliente '{cliente}'!")
    except Exception as e:
        messagebox.showerror("Erro", f"Não foi possível salvar a planilha. Erro: {str(e)}")

# Função para adicionar um item à lista e limpar os campos de entrada
def adicionar_item():
    instrumento = instrumento_entry.get()
    resolucao = resolucao_entry.get()
    capacidade = capacidade_entry.get()
    codigo = codigo_entry.get()
    modelo = modelo_entry.get()
    fabricante = fabricante_entry.get()
    cliente = cliente_entry.get()
    manutencao = manutencao_entry.get()
    valor_total_str = valor_total_entry.get()
    protocolo = protocolo_entry.get()
    status = status_var.get()
    if not status or status.lower() == "none":
        status = ""

    # Verificar se o valor_total é válido
    try:
        valor_total = float(valor_total_str) if valor_total_str else 0.0
    except ValueError:
        messagebox.showerror("Erro", "Valor Unitário inválido. Por favor, insira um número válido.")
        return

    item = ItemOrcamento(instrumento, resolucao, capacidade, codigo, modelo, fabricante, cliente, manutencao, valor_total, protocolo, status)
    itens.append(item)

    # Obter o próximo número de item
    next_number = len(itens)
    
    # Adicionar o item à Treeview com o número
    tree.insert("", "end", values=(next_number, instrumento, resolucao, capacidade, codigo, modelo, fabricante, protocolo, cliente, manutencao, f"{valor_total:.2f}", status))

    # Limpar os campos de entrada
    instrumento_entry.delete(0, tk.END)
    resolucao_entry.delete(0, tk.END)
    capacidade_entry.delete(0, tk.END)
    codigo_entry.delete(0, tk.END)
    modelo_entry.delete(0, tk.END)
    fabricante_entry.delete(0, tk.END)
    cliente_entry.delete(0, tk.END)
    manutencao_entry.delete(0, tk.END)
    valor_total_entry.delete(0, tk.END)
    protocolo_entry.delete(0, tk.END)
    status_var.set("")  # Reset status dropdown

# Função para deletar os itens selecionados
def deletar_item(event=None):
    selected_items = tree.selection()
    if selected_items:
        # Confirmar a exclusão
        if messagebox.askyesno("Confirmar exclusão", f"Tem certeza que deseja excluir {len(selected_items)} item(s)?"):
            # Ordenar os índices em ordem decrescente para evitar problemas com a remoção
            indices = sorted([tree.index(item) for item in selected_items], reverse=True)
            for index in indices:
                tree.delete(selected_items[indices.index(index)])
                del itens[index]
    else:
        messagebox.showwarning("Seleção vazia", "Por favor, selecione um ou mais itens para deletar.")

# Função para editar o item selecionado
def editar_item(event=None):
    selected_item = tree.selection()
    if selected_item:
        item_index = tree.index(selected_item[0])
        instrumento = tree.item(selected_item, 'values')[1]
        resolucao = tree.item(selected_item, 'values')[2]
        capacidade = tree.item(selected_item, 'values')[3]
        codigo = tree.item(selected_item, 'values')[4]
        modelo = tree.item(selected_item, 'values')[5]
        fabricante = tree.item(selected_item, 'values')[6]
        protocolo = tree.item(selected_item, 'values')[7]
        cliente = tree.item(selected_item, 'values')[8]
        manutencao = tree.item(selected_item, 'values')[9]
        valor_total = float(tree.item(selected_item, 'values')[10])
        status = tree.item(selected_item, 'values')[11]

        instrumento_entry.delete(0, tk.END)
        instrumento_entry.insert(0, instrumento)
        resolucao_entry.delete(0, tk.END)
        resolucao_entry.insert(0, resolucao)
        capacidade_entry.delete(0, tk.END)
        capacidade_entry.insert(0, capacidade)
        codigo_entry.delete(0, tk.END)
        codigo_entry.insert(0, codigo)
        modelo_entry.delete(0, tk.END)
        modelo_entry.insert(0, modelo)
        fabricante_entry.delete(0, tk.END)
        fabricante_entry.insert(0, fabricante)
        protocolo_entry.delete(0, tk.END)
        protocolo_entry.insert(0, protocolo)
        cliente_entry.delete(0, tk.END)
        cliente_entry.insert(0, cliente)
        manutencao_entry.delete(0, tk.END)
        manutencao_entry.insert(0, manutencao)
        valor_total_entry.delete(0, tk.END)
        valor_total_entry.insert(0, f"{valor_total:.2f}")
        status_var.set(status)

        # Deletar o item da lista
        del itens[item_index]
        tree.delete(selected_item)

# Função para salvar as edições do item após a edição na interface
def salvar_edicao():
    instrumento = instrumento_entry.get()
    resolucao = resolucao_entry.get()
    capacidade = capacidade_entry.get()
    codigo = codigo_entry.get()
    modelo = modelo_entry.get()
    fabricante = fabricante_entry.get()
    cliente = cliente_entry.get()
    manutencao = manutencao_entry.get()
    valor_total = float(valor_total_entry.get())
    protocolo = protocolo_entry.get()
    status = status_var.get()
    if not status or status.lower() == "none":
        status = ""

    item = ItemOrcamento(instrumento, resolucao, capacidade, codigo, modelo, fabricante, cliente, manutencao, valor_total, protocolo, status)
    itens.append(item)

    tree.insert("", "end", values=(len(itens), instrumento, resolucao, capacidade, codigo, modelo, fabricante, protocolo, cliente, manutencao, valor_total, status))

    instrumento_entry.delete(0, tk.END)
    resolucao_entry.delete(0, tk.END)
    capacidade_entry.delete(0, tk.END)
    codigo_entry.delete(0, tk.END)
    modelo_entry.delete(0, tk.END)
    fabricante_entry.delete(0, tk.END)
    cliente_entry.delete(0, tk.END)
    manutencao_entry.delete(0, tk.END)
    valor_total_entry.delete(0, tk.END)
    protocolo_entry.delete(0, tk.END)
    status_var.set("")

# Função para gerar a planilha com os itens adicionados
def gerar_planilha():
    if not itens:
        messagebox.showwarning("Lista vazia", "Adicione pelo menos um item antes de gerar a planilha.")
        return

    nome_arquivo = planilha_entry.get().strip()
    if not nome_arquivo:
        messagebox.showwarning("Campo vazio", "Por favor, insira um nome para o arquivo Excel.")
        return

    if not nome_arquivo.endswith(".xlsx"):
        nome_arquivo += ".xlsx"
    
    numero_orcamento = numero_orcamento_entry.get()
    data = data_entry.get()
    
    criar_planilha_orcamento(itens, nome_arquivo, numero_orcamento, data)

# Função para importar uma planilha Excel
def importar_planilha():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if not file_path:
        return

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Limpar a lista de itens e a Treeview
    itens.clear()
    for row in tree.get_children():
        tree.delete(row)

    # Ler os dados da planilha e adicionar à lista de itens
    for index, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=1):
        if any(row):
            # Check if the first cell is not "Valor Total" (as string or as part of a string)
            if not (isinstance(row[0], str) and "Valor Total" in row[0]):
                try:
                    # Converter os valores para o tipo correto
                    converted_row = [
                        str(row[1]) if len(row) > 1 and row[1] is not None else "",  # instrumento
                        str(row[2]) if len(row) > 2 and row[2] is not None else "",  # resolucao
                        str(row[3]) if len(row) > 3 and row[3] is not None else "",  # capacidade
                        str(row[4]) if len(row) > 4 and row[4] is not None else "",  # codigo
                        str(row[5]) if len(row) > 5 and row[5] is not None else "",  # modelo
                        str(row[6]) if len(row) > 6 and row[6] is not None else "",  # fabricante
                        str(row[8]) if len(row) > 8 and row[8] is not None else "",  # cliente
                        str(row[9]) if len(row) > 9 and row[9] is not None else "",  # manutencao
                        float(row[10]) if len(row) > 10 and row[10] is not None else 0.0,  # valor_total
                        str(row[7]) if len(row) > 7 and row[7] is not None else "",  # protocolo
                        str(row[11]) if len(row) > 11 and row[11] is not None else ""  # status
                    ]
                    
                    # Tratar o status
                    status = converted_row[10]
                    if not status or status.lower() == "none":
                        status = ""
                    
                    # Criar o item
                    item = ItemOrcamento(*converted_row)
                    itens.append(item)
                    
                    # Inserir na Treeview
                    tree.insert("", "end", values=(index, *converted_row))
                except ValueError as e:
                    print(f"Erro ao converter valor na linha {index}: {e}")
                    continue  # Pula para a próxima linha em caso de erro

    messagebox.showinfo("Sucesso", f"Planilha '{file_path}' importada com sucesso!")

# Função para mostrar a janela "Sobre"
def show_about():
    about_window = tk.Toplevel(root)
    about_window.title("Sobre")
    about_window.geometry("300x100")
    about_window.resizable(False, False)

    tk.Label(about_window, text="Sistema feito por Gabriel Businari", pady=10).pack()
    tk.Label(about_window, text="©2024 Todos os direitos reservados").pack()

# Função para obter ou criar a pasta do cliente
def obter_pasta_cliente(cliente):
    pasta_cliente = os.path.join("Clientes", cliente)
    if not os.path.exists(pasta_cliente):
        os.makedirs(pasta_cliente)
    return pasta_cliente

# Configuração da interface gráfica
root = tk.Tk()
root.title("Sistema de Orçamento")

# Lista para armazenar os itens do orçamento
itens = []

# Labels e Entradas para os itens
tk.Label(root, text="Instrumento").grid(row=0, column=0, padx=5, pady=5, sticky="ew")
instrumento_entry = tk.Entry(root)
instrumento_entry.grid(row=0, column=1, padx=5, pady=5, sticky="ew")

tk.Label(root, text="Resolução (mm)").grid(row=1, column=0, padx=5, pady=5, sticky="ew")
resolucao_entry = tk.Entry(root)
resolucao_entry.grid(row=1, column=1, padx=5, pady=5, sticky="ew")

tk.Label(root, text="Capacidade (mm)").grid(row=2, column=0, padx=5, pady=5, sticky="ew")
capacidade_entry = tk.Entry(root)
capacidade_entry.grid(row=2, column=1, padx=5, pady=5, sticky="ew")

tk.Label(root, text="Código").grid(row=3, column=0, padx=5, pady=5, sticky="ew")
codigo_entry = tk.Entry(root)
codigo_entry.grid(row=3, column=1, padx=5, pady=5, sticky="ew")

tk.Label(root, text="Modelo").grid(row=4, column=0, padx=5, pady=5, sticky="ew")
modelo_entry = tk.Entry(root)
modelo_entry.grid(row=4, column=1, padx=5, pady=5, sticky="ew")

tk.Label(root, text="Fabricante").grid(row=5, column=0, padx=5, pady=5, sticky="ew")
fabricante_entry = tk.Entry(root)
fabricante_entry.grid(row=5, column=1, padx=5, pady=5, sticky="ew")

tk.Label(root, text="Protocolo").grid(row=6, column=0, padx=5, pady=5, sticky="ew")
protocolo_entry = tk.Entry(root)
protocolo_entry.grid(row=6, column=1, padx=5, pady=5, sticky="ew")

tk.Label(root, text="Cliente").grid(row=7, column=0, padx=5, pady=5, sticky="ew")
cliente_entry = tk.Entry(root)
cliente_entry.grid(row=7, column=1, padx=5, pady=5, sticky="ew")

tk.Label(root, text="Manutenção").grid(row=8, column=0, padx=5, pady=5, sticky="ew")
manutencao_entry = tk.Entry(root)
manutencao_entry.grid(row=8, column=1, padx=5, pady=5, sticky="ew")

tk.Label(root, text="Valor Unitário").grid(row=9, column=0, padx=5, pady=5, sticky="ew")
valor_total_entry = tk.Entry(root)
valor_total_entry.grid(row=9, column=1, padx=5, pady=5, sticky="ew")

# Labels e Entradas para o número do orçamento e data
tk.Label(root, text="Número do Orçamento").grid(row=10, column=0, padx=5, pady=5, sticky="ew")
numero_orcamento_entry = tk.Entry(root)
numero_orcamento_entry.grid(row=10, column=1, padx=5, pady=5, sticky="ew")

tk.Label(root, text="Data").grid(row=11, column=0, padx=5, pady=5, sticky="ew")
data_entry = tk.Entry(root)
data_entry.grid(row=11, column=1, padx=5, pady=5, sticky="ew")

# Dropdown para o status
tk.Label(root, text="Status").grid(row=12, column=0, padx=5, pady=5, sticky="ew")
status_var = tk.StringVar(root)
status_dropdown = ttk.Combobox(root, textvariable=status_var, values=["", "Aprovado", "Reprovado"], state="readonly")  # Definido como readonly
status_dropdown.grid(row=12, column=1, padx=5, pady=5, sticky="ew")

# Labels e Entradas para o nome dos arquivos
tk.Label(root, text="Nome do arquivo Excel").grid(row=13, column=0, padx=5, pady=5, sticky="ew")
planilha_entry = tk.Entry(root)
planilha_entry.grid(row=13, column=1, padx=5, pady=5, sticky="ew")

# Botão para adicionar item
add_button = tk.Button(root, text="Adicionar Item", command=adicionar_item)
add_button.grid(row=14, column=0, columnspan=2, padx=5, pady=5, sticky="ew")

# Botão para editar item
edit_button = tk.Button(root, text="Editar Item", command=editar_item)
edit_button.grid(row=15, column=0, columnspan=2, padx=5, pady=5, sticky="ew")

# Treeview para exibir os itens adicionados
columns = ("numero", "instrumento", "resolucao", "capacidade", "codigo", "modelo", "fabricante", "protocolo", "cliente", "manutencao", "valor_total", "status")
tree = ttk.Treeview(root, columns=columns, show="headings")
tree.heading("numero", text="Item")
tree.heading("instrumento", text="Instrumento")
tree.heading("resolucao", text="Resolução (mm)")
tree.heading("capacidade", text="Capacidade (mm)")
tree.heading("codigo", text="Código")
tree.heading("modelo", text="Modelo")
tree.heading("fabricante", text="Fabricante")
tree.heading("protocolo", text="Protocolo")
tree.heading("cliente", text="Cliente")
tree.heading("manutencao", text="Manutenção")
tree.heading("valor_total", text="Valor Total")
tree.heading("status", text="Status")
tree.grid(row=16, column=0, columnspan=2, padx=5, pady=5, sticky="nsew")

# Adicionar scrollbar
scrollbar = ttk.Scrollbar(root, orient="vertical", command=tree.yview)
tree.configure(yscrollcommand=scrollbar.set)
scrollbar.grid(row=16, column=2, sticky="ns")

# Bind da tecla Del para deletar item
root.bind("<Delete>", deletar_item)

# Botão para gerar a planilha
generate_button = tk.Button(root, text="Gerar Planilha", command=gerar_planilha)
generate_button.grid(row=17, column=0, columnspan=2, padx=5, pady=5, sticky="ew")

# Botão para importar planilha
import_button = tk.Button(root, text="Importar Planilha", command=importar_planilha)
import_button.grid(row=18, column=0, columnspan=2, padx=5, pady=5, sticky="ew")

# Criação da barra de menu
menu_bar = Menu(root)
root.config(menu=menu_bar)

# Criação do menu "Sobre"
about_menu = Menu(menu_bar, tearoff=0)
menu_bar.add_cascade(label="Sobre", menu=about_menu)
about_menu.add_command(label="Informações", command=show_about)

# Configuração das colunas e linhas para redimensionamento
root.grid_columnconfigure(0, weight=1)
root.grid_columnconfigure(1, weight=1)
root.grid_rowconfigure(16, weight=1)

root.mainloop()
